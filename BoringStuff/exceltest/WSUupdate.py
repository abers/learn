#! python3

import openpyxl
import os
import shutil

print('Moving to correct directory')
os.chdir('E:\\Learning\\Python\\BoringStuff\\exceltest')

print('Backing up files')
try:
    shutil.copy2('WSU_Attributes_2.xlsx', 'WSU_Attributes_3.xlsx')
except FileNotFoundError:
    pass
shutil.copy2('WSU_Attributes.xlsx', 'WSU_Attributes_2.xlsx')


# Change numbers to correct format
print('Loading...')

WSU = openpyxl.load_workbook('Classification Sheet - WSU.xlsx')
WSUdata = WSU.get_sheet_by_name('WSU')

'''
WSUNew = openpyxl.load_workbook('WSU_Attributes.xlsx')
WSUNewdata = WSUNew.get_sheet_by_name('WSU')
'''

for rowNum in range(2, WSUdata.max_row + 1):
    try:
        print(WSUdata.cell(row=1, column=5).value)
        WSUdata.cell(row=rowNum, column=5).value = int(WSUdata.cell(row=rowNum, column=5).value)
        print('Change OK')
    except ValueError:
        print('Unassigned or Not Applicable')
        continue
    try:
        print(WSUdata.cell(row=1, column=101).value)
        WSUdata.cell(row=rowNum, column=101).value = int(WSUdata.cell(row=rowNum, column=101).value)
        print('Change OK')
    except ValueError:
        print('Unassigned or Not Applicable')
        continue
    try:
        print(WSUdata.cell(row=1, column=198).value)
        WSUdata.cell(row=rowNum, column=198).value = int(WSUdata.cell(row=rowNum, column=198).value)
        print('Change OK')
    except ValueError:
        print('Unassigned or Not Applicable')
        continue
WSU.save('Classification Sheet - WSU.xlsx')
'''
# Copy data to new sheet
for rowNum in range(2, WSUdata.max_row + 1):
    for colNum in range(1, WSUdata.max_column):
        WSUNewdata.cell(row=rowNum, column=colNum).value = WSUdata.cell(row=rowNum, column=colNum).value
    print(str(rowNum) + ' complete')

print('Saving...')
WSUNew.save('WSU_Attributes.xlsx')
print('Complete')
'''
